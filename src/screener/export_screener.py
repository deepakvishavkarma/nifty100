"""screener_output.xlsx generator - 6 sheets, colour-coded (Sprint 3, Day 17)."""
from __future__ import annotations
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)

DISPLAY_COLS = [
    "company_id", "company_name", "broad_sector", "return_on_equity_pct",
    "return_on_capital_employed_pct", "net_profit_margin_pct", "debt_to_equity",
    "interest_coverage", "icr_label", "asset_turnover", "free_cash_flow_cr",
    "fcf_yield_pct", "revenue_cagr_5yr", "pat_cagr_5yr", "eps_cagr_5yr",
    "pe_ratio", "pb_ratio", "dividend_yield_pct", "dividend_payout_ratio_pct",
    "market_cap_crore", "composite_quality_score",
]

COL_LABELS = {
    "company_id": "Ticker", "company_name": "Company", "broad_sector": "Sector",
    "return_on_equity_pct": "ROE %", "return_on_capital_employed_pct": "ROCE %",
    "net_profit_margin_pct": "NPM %", "debt_to_equity": "D/E",
    "interest_coverage": "ICR", "icr_label": "ICR Label", "asset_turnover": "Asset TO",
    "free_cash_flow_cr": "FCF (Cr)", "fcf_yield_pct": "FCF Yield %",
    "revenue_cagr_5yr": "Rev CAGR 5yr %", "pat_cagr_5yr": "PAT CAGR 5yr %",
    "eps_cagr_5yr": "EPS CAGR 5yr %", "pe_ratio": "P/E", "pb_ratio": "P/B",
    "dividend_yield_pct": "Div Yield %", "dividend_payout_ratio_pct": "Payout %",
    "market_cap_crore": "Mkt Cap (Cr)", "composite_quality_score": "Quality Score",
}


def _passes_threshold(value, metric: str, filters: dict):
    """True/False if this metric has a threshold in the preset's filters, else None."""
    if metric not in filters or value is None or pd.isna(value):
        return None
    cond = filters[metric]
    if "min" in cond:
        return value >= cond["min"]
    if "max" in cond:
        return value <= cond["max"]
    if "eq" in cond:
        return value == cond["eq"]
    return None


def write_preset_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, preset_label: str, filters: dict):
    ws = wb.create_sheet(sheet_name[:31])
    ws["A1"] = f"{preset_label} — {len(df)} companies"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DISPLAY_COLS))

    header_row = 3
    for c, col in enumerate(DISPLAY_COLS, start=1):
        cell = ws.cell(row=header_row, column=c, value=COL_LABELS.get(col, col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for r, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        for c, col in enumerate(DISPLAY_COLS, start=1):
            value = row.get(col)
            if pd.isna(value):
                value = None
            elif isinstance(value, float):
                value = round(value, 2)
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = BODY_FONT
            passes = _passes_threshold(row.get(col), col, filters)
            if passes is True:
                cell.fill = GREEN
            elif passes is False:
                cell.fill = RED

    for c, col in enumerate(DISPLAY_COLS, start=1):
        width = max(len(COL_LABELS.get(col, col)) + 2, 12)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def generate_screener_workbook(results: dict, config: dict, out_path: Path):
    wb = Workbook()
    wb.remove(wb.active)
    for key, df in results.items():
        preset = config["presets"][key]
        write_preset_sheet(wb, preset["label"], df, preset["label"], preset["filters"])
    wb.save(out_path)
