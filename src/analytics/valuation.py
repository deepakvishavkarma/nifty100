"""Valuation module (Sprint 4, Day 26).

Uses market_cap.xlsx (loaded into SQLite in Sprint 1) + financial_ratios
(Sprint 2) to compute FCF yield, sector-relative P/E flags, and export
valuation_summary.xlsx + valuation_flags.csv.
"""
from __future__ import annotations
import sys
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

ROOT = Path(__file__).resolve().parent.parent.parent
DB_PATH = ROOT / "data" / "nifty100.db"
OUT_DIR = ROOT / "output"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)


def _latest_common_year(df: pd.DataFrame) -> str:
    counts = df["year"].value_counts()
    threshold = counts.max() * 0.5
    return counts[counts >= threshold].index.max()


def build_valuation_table() -> pd.DataFrame:
    conn = sqlite3.connect(DB_PATH)
    mc = pd.read_sql("SELECT * FROM market_cap", conn)
    fr = pd.read_sql("SELECT company_id, year, free_cash_flow_cr FROM financial_ratios", conn)
    sectors = pd.read_sql("SELECT company_id, broad_sector FROM sectors", conn)
    companies = pd.read_sql("SELECT id AS company_id, company_name FROM companies", conn)
    conn.close()

    latest_year = _latest_common_year(mc)
    latest_mc = mc[mc["year"] == latest_year].copy()
    latest_fcf = fr[fr["year"] == latest_year][["company_id", "free_cash_flow_cr"]]

    df = (latest_mc
          .merge(latest_fcf, on="company_id", how="left")
          .merge(sectors, on="company_id", how="left")
          .merge(companies, on="company_id", how="left"))

    # FCF Yield = FCF / market_cap_crore x 100
    df["fcf_yield_pct"] = df.apply(
        lambda r: (r["free_cash_flow_cr"] / r["market_cap_crore"] * 100)
        if pd.notna(r["free_cash_flow_cr"]) and pd.notna(r["market_cap_crore"]) and r["market_cap_crore"] != 0
        else None, axis=1)

    # 5yr median P/E per company (own history, not sector)
    pe_hist = mc[["company_id", "year", "pe_ratio"]]
    five_yr_median = pe_hist.groupby("company_id")["pe_ratio"].median().rename("pe_5yr_median")
    df = df.merge(five_yr_median, on="company_id", how="left")

    # Sector median P/E (cross-sectional, latest year)
    sector_median_pe = df.groupby("broad_sector")["pe_ratio"].median().rename("sector_median_pe")
    df = df.merge(sector_median_pe, on="broad_sector", how="left")

    df["pe_vs_sector_median_pct"] = df.apply(
        lambda r: ((r["pe_ratio"] - r["sector_median_pe"]) / r["sector_median_pe"] * 100)
        if pd.notna(r["pe_ratio"]) and pd.notna(r["sector_median_pe"]) and r["sector_median_pe"] != 0
        else None, axis=1)

    def flag(row):
        pe, smed = row["pe_ratio"], row["sector_median_pe"]
        if pd.isna(pe) or pd.isna(smed) or smed == 0:
            return None
        if pe > smed * 1.5:
            return "Caution"
        if pe < smed * 0.7:
            return "Discount"
        return "Fair"

    df["flag"] = df.apply(flag, axis=1)
    df["year"] = latest_year

    return df[[
        "company_id", "company_name", "broad_sector", "pe_ratio", "pb_ratio", "ev_ebitda",
        "fcf_yield_pct", "pe_5yr_median", "sector_median_pe", "pe_vs_sector_median_pct",
        "flag", "dividend_yield_pct", "market_cap_crore", "year",
    ]].rename(columns={"broad_sector": "sector", "pe_ratio": "PE", "pb_ratio": "PB",
                        "ev_ebitda": "EV_EBITDA", "pe_5yr_median": "5yr_median_PE"})


def export_valuation_summary(df: pd.DataFrame, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation Summary"

    display_cols = ["company_id", "company_name", "sector", "PE", "PB", "EV_EBITDA",
                     "fcf_yield_pct", "5yr_median_PE", "pe_vs_sector_median_pct", "flag"]
    labels = {"company_id": "Ticker", "company_name": "Company", "sector": "Sector",
              "PE": "P/E", "PB": "P/B", "EV_EBITDA": "EV/EBITDA", "fcf_yield_pct": "FCF Yield %",
              "5yr_median_PE": "5yr Median P/E", "pe_vs_sector_median_pct": "P/E vs Sector %",
              "flag": "Flag"}
    flag_fill = {"Caution": YELLOW, "Discount": GREEN, "Fair": None}

    for c, col in enumerate(display_cols, start=1):
        cell = ws.cell(row=1, column=c, value=labels[col])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for r, (_, row) in enumerate(df.sort_values("company_id").iterrows(), start=2):
        for c, col in enumerate(display_cols, start=1):
            value = row.get(col)
            if pd.isna(value):
                value = None
            elif isinstance(value, float):
                value = round(value, 2)
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = BODY_FONT
            if col == "flag" and value in flag_fill and flag_fill[value]:
                cell.fill = flag_fill[value]

    for c, col in enumerate(display_cols, start=1):
        ws.column_dimensions[get_column_letter(c)].width = max(len(labels[col]) + 2, 14)
    ws.freeze_panes = "A2"
    wb.save(out_path)


def export_valuation_flags(df: pd.DataFrame, out_path: Path):
    flagged = df[df["flag"].isin(["Caution", "Discount"])].copy()
    flagged = flagged.sort_values(["flag", "pe_vs_sector_median_pct"], ascending=[True, False])
    flagged.to_csv(out_path, index=False)
    return flagged


def main():
    OUT_DIR.mkdir(exist_ok=True)
    df = build_valuation_table()
    export_valuation_summary(df, OUT_DIR / "valuation_summary.xlsx")
    flagged = export_valuation_flags(df, OUT_DIR / "valuation_flags.csv")
    print(f"valuation_summary.xlsx: {len(df)} rows")
    print(f"valuation_flags.csv: {len(flagged)} rows")
    print(df["flag"].value_counts())


if __name__ == "__main__":
    main()
