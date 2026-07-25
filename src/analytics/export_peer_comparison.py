"""peer_comparison.xlsx generator — 11 sheets, one per peer group (Sprint 3, Day 20)."""
from __future__ import annotations
import sys
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
from src.screener.engine import load_universe
from src.analytics.peer import PEER_METRICS

ROOT = Path(__file__).resolve().parent.parent.parent
DB_PATH = ROOT / "data" / "nifty100.db"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)

METRIC_LABELS = {
    "roe": "ROE %", "roce": "ROCE %", "net_profit_margin": "NPM %",
    "debt_to_equity": "D/E", "free_cash_flow": "FCF (Cr)", "pat_cagr_5yr": "PAT CAGR 5yr %",
    "revenue_cagr_5yr": "Rev CAGR 5yr %", "eps_cagr_5yr": "EPS CAGR 5yr %",
    "interest_coverage": "ICR", "asset_turnover": "Asset Turnover",
}
METRIC_KEYS = list(PEER_METRICS.keys())


def _pct_fill(pct: float | None):
    if pct is None:
        return None
    if pct >= 0.75:
        return GREEN
    if pct <= 0.25:
        return RED
    return YELLOW


def write_group_sheet(wb: Workbook, group_name: str, group_df: pd.DataFrame,
                       pct_df: pd.DataFrame, benchmark_id: str | None):
    ws = wb.create_sheet(group_name[:31])
    ws["A1"] = f"{group_name} — {len(group_df)} companies"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)

    # header: company_id, company_name, then metric | metric_pctile pairs
    headers = ["Ticker", "Company"]
    for m in METRIC_KEYS:
        headers += [METRIC_LABELS[m], f"{METRIC_LABELS[m]} %ile"]
    n_cols = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    header_row = 3
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    r = header_row + 1
    for _, row in group_df.sort_values("company_id").iterrows():
        cid = row["company_id"]
        ws.cell(row=r, column=1, value=cid)
        ws.cell(row=r, column=2, value=row.get("company_name"))
        c = 3
        for m in METRIC_KEYS:
            col = PEER_METRICS[m][0]
            raw = row.get(col)
            sub = pct_df[(pct_df["company_id"] == cid) & (pct_df["metric"] == m)]
            pct = sub["percentile_rank"].iloc[0] if len(sub) else None

            vcell = ws.cell(row=r, column=c, value=None if pd.isna(raw) else round(float(raw), 2))
            pcell = ws.cell(row=r, column=c + 1, value=None if pct is None else round(pct * 100, 1))
            fill = _pct_fill(pct)
            if fill:
                pcell.fill = fill
            c += 2

        is_bench = benchmark_id is not None and cid == benchmark_id
        row_font = BOLD_FONT if is_bench else BODY_FONT
        for cc in range(1, n_cols + 1):
            ws.cell(row=r, column=cc).font = row_font
        if is_bench:
            # gold background on the label columns; percentile columns keep
            # their traffic-light colour so the ranking stays readable
            ws.cell(row=r, column=1).fill = GOLD
            ws.cell(row=r, column=2).fill = GOLD
        r += 1

    # summary row: peer group median for each metric
    ws.cell(row=r, column=1, value="Peer Group Median").font = BOLD_FONT
    c = 3
    for m in METRIC_KEYS:
        col = PEER_METRICS[m][0]
        median = group_df[col].median()
        cell = ws.cell(row=r, column=c, value=None if pd.isna(median) else round(float(median), 2))
        cell.font = BOLD_FONT
        c += 2
    r += 1

    for c in range(1, n_cols + 1):
        letter = get_column_letter(c)
        ws.column_dimensions[letter].width = 14 if c > 2 else (10 if c == 1 else 26)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=3)


def generate_peer_workbook(out_path: Path):
    conn = sqlite3.connect(DB_PATH)
    universe = load_universe(conn)
    peer_groups = pd.read_sql("SELECT peer_group_name, company_id, is_benchmark FROM peer_groups", conn)
    pct_df = pd.read_sql("SELECT * FROM peer_percentiles", conn)
    conn.close()

    wb = Workbook()
    wb.remove(wb.active)

    for group_name, members in peer_groups.groupby("peer_group_name"):
        group_universe = universe[universe["company_id"].isin(members["company_id"])].copy()
        bench_row = members[members["is_benchmark"] == 1]
        benchmark_id = bench_row["company_id"].iloc[0] if len(bench_row) else None
        write_group_sheet(wb, group_name, group_universe, pct_df, benchmark_id)

    wb.save(out_path)
    print(f"peer_comparison.xlsx written: {len(peer_groups['peer_group_name'].unique())} sheets")


if __name__ == "__main__":
    generate_peer_workbook(ROOT / "output" / "peer_comparison.xlsx")
